"""無料データソースからの取得。すべて標準ライブラリのみ・APIキー不要。

各関数は取得失敗時に None を返し、エラーは errors リストに積む。
一部が落ちても天気予報全体は残りのデータで組み立てる（縮退運転）。
"""
from __future__ import annotations

import csv
import io
import json
import time
import urllib.request
from datetime import datetime, timezone

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)
TIMEOUT = 45

GVZ_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=GVZCLS"
STOOQ_URL = "https://stooq.com/q/d/l/?s=xauusd&i=d"
COT_URL = (
    "https://publicreporting.cftc.gov/resource/72hh-3qpy.json"
    "?cftc_contract_market_code=088691"
    "&$order=report_date_as_yyyy_mm_dd%20DESC&$limit=1100"
)
# 2026年初頭のサイト刷新で旧CSVは廃止。現行は公式API（XLSX全履歴 / JSON当日値）
GLD_XLSX_URL = (
    "https://api.spdrgoldshares.com/api/v1/historical-archive"
    "?product=gld&exchange=NYSE&lang=en"
)
GLD_JSON_URL = (
    "https://api.spdrgoldshares.com/api/v1/data?product=gld&exchange=nyse&lang=en"
)
FF_CAL_URL = "https://nfs.faireconomy.media/ff_calendar_thisweek.json"


def _get(url: str, tries: int = 3, headers: dict | None = None) -> bytes:
    last: Exception | None = None
    hdrs = {"User-Agent": UA, **(headers or {})}
    for i in range(tries):
        try:
            req = urllib.request.Request(url, headers=hdrs)
            with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
                return resp.read()
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(2 * (i + 1))
    raise last  # type: ignore[misc]


def _head(text: str) -> str:
    """エラーメッセージ用に応答の先頭を短く返す（改行・非印字文字は除去）。"""
    clean = "".join(c for c in text[:120] if c.isprintable())
    return clean[:80]


def _yahoo_chart(symbol: str, range_: str, interval: str) -> dict:
    """Yahoo Finance chart API。キー不要・サーバーからのアクセスに比較的寛容。

    戻り値: {"dates": [ISO日付], "open": [...], "high": [...], "low": [...], "close": [...]}
    欠損(None)は落とす。
    """
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
        f"?range={range_}&interval={interval}"
    )
    data = json.loads(_get(url).decode("utf-8", "replace"))
    result = data["chart"]["result"][0]
    ts = result["timestamp"]
    q = result["indicators"]["quote"][0]
    out: dict = {
        "dates": [], "ts": [], "open": [], "high": [], "low": [], "close": []
    }
    for i, t in enumerate(ts):
        c = q["close"][i]
        if c is None:
            continue
        out["dates"].append(
            datetime.fromtimestamp(t, tz=timezone.utc).strftime("%Y-%m-%d")
        )
        out["ts"].append(int(t))
        out["close"].append(c)
        out["open"].append(q["open"][i] if q["open"][i] is not None else c)
        out["high"].append(q["high"][i] if q["high"][i] is not None else c)
        out["low"].append(q["low"][i] if q["low"][i] is not None else c)
    if not out["close"]:
        raise ValueError(f"no data for {symbol}")
    return out


def fetch_gvz(errors: list[str]) -> dict | None:
    """CBOE GVZ（金のインプライドボラ指数）。

    第一候補: Yahoo ^GVZ（10年分・サーバーからでも通りやすい）
    第二候補: FRED CSV（実行元IPによってはタイムアウトする）
    """
    try:
        ch = _yahoo_chart("%5EGVZ", "10y", "1d")
        series = list(zip(ch["dates"], ch["close"]))
        return {"series": series, "date": series[-1][0], "value": series[-1][1]}
    except Exception as e_yahoo:  # noqa: BLE001
        try:
            text = _get(GVZ_URL).decode("utf-8", "replace")
            rows = list(csv.reader(io.StringIO(text)))
            series = [
                (r[0], float(r[1]))
                for r in rows[1:]
                if len(r) >= 2 and r[1] not in (".", "")
            ]
            if not series:
                raise ValueError("empty series")
            return {"series": series, "date": series[-1][0], "value": series[-1][1]}
        except Exception as e:  # noqa: BLE001
            errors.append(
                f"GVZ取得失敗: yahoo={type(e_yahoo).__name__}:{e_yahoo} / "
                f"fred={type(e).__name__}:{e}"
            )
            return None


def fetch_price(errors: list[str]) -> dict | None:
    """金価格の日足。前日高値安値と直近終値に使う。

    第一候補: Yahoo XAUUSD=X（スポット）、駄目なら GC=F（COMEX先物・ほぼ同水準）。
    """
    last_err: Exception | None = None
    for symbol in ("XAUUSD%3DX", "GC%3DF"):
        try:
            ch = _yahoo_chart(symbol, "5d", "1d")
            i = len(ch["close"]) - 1
            # 末尾行は「進行中の当日」のことがある(朝実行時は数時間分しかない)。
            # 前日高安には、直近の「完結した1日」= 末尾より前で最も新しい行を使う
            j = i - 1 if i >= 1 else i
            return {
                "date": ch["dates"][j],
                "open": ch["open"][j],
                "high": ch["high"][j],
                "low": ch["low"][j],
                "close": ch["close"][i],  # 現在値は最新
                "symbol": "spot" if "XAUUSD" in symbol else "futures(GC=F)",
            }
        except Exception as e:  # noqa: BLE001
            last_err = e
    errors.append(f"価格取得失敗: {type(last_err).__name__}: {last_err}")
    return None


def _find_key(row: dict, *needles: str) -> str | None:
    for k in row:
        lk = k.lower()
        if all(n in lk for n in needles):
            return k
    return None


def fetch_cot(errors: list[str]) -> dict | None:
    """CFTC COT（金・Disaggregated先物、週次）。Socrata API、認証不要。

    フィールド名はSocrata側の命名揺れに備えて部分一致で解決する。
    """
    try:
        data = json.loads(_get(COT_URL).decode("utf-8", "replace"))
        if not data:
            raise ValueError("empty response")
        r0 = data[0]
        k_long = _find_key(r0, "m_money", "long")
        k_short = _find_key(r0, "m_money", "short")
        k_oi = _find_key(r0, "open_interest_all") or _find_key(r0, "open_interest")
        k_date = _find_key(r0, "report_date")
        if not (k_long and k_short and k_oi and k_date):
            raise ValueError(f"fields not found in keys: {list(r0)[:10]}...")
        hist = []
        for r in data:
            try:
                net = float(r[k_long]) - float(r[k_short])
                oi = float(r[k_oi])
                if oi > 0:
                    hist.append((r[k_date][:10], net / oi))
            except (KeyError, ValueError, TypeError):
                continue
        if not hist:
            raise ValueError("no parsable rows")
        latest = hist[0]
        values = [v for _, v in hist]
        rank = sum(1 for v in values if v <= latest[1]) / len(values)
        return {
            "date": latest[0],
            "net_over_oi": latest[1],
            "percentile": round(rank * 100),
            "n_weeks": len(values),
        }
    except Exception as e:  # noqa: BLE001
        errors.append(f"COT取得失敗: {type(e).__name__}: {e}")
        return None


def fetch_gld(errors: list[str]) -> dict | None:
    """GLD（最大の金ETF）の保有トン数。

    第一候補: 公式APIのXLSX全履歴（2004年〜、Tonnes列=10列目）。
    ブラウザ相当のOrigin/Referer必須（無いとPDFにすり替えられる）。
    第二候補: 公式APIのJSON当日値（履歴なし→前週比は出せない）。
    """
    hdrs = {
        "Accept": (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet,*/*"
        ),
        "Origin": "https://www.spdrgoldshares.com",
        "Referer": "https://www.spdrgoldshares.com/usa/historical-data/",
    }
    try:
        raw = _get(GLD_XLSX_URL, headers=hdrs)
        if raw[:4] == b"%PDF":
            raise ValueError("PDFが返却された（ヘッダー不足の可能性）")
        from openpyxl import load_workbook  # Actionsでのみ必要な依存

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb["US GLD Historical Archive"]
        series: list[tuple[str, float]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None or len(row) < 10:
                continue
            tonnes = row[9]
            if isinstance(tonnes, (int, float)):  # 休場日は文字列なので除外
                series.append((str(row[0])[:11], float(tonnes)))
        if len(series) < 6:
            raise ValueError(f"rows too few: {len(series)}")
        latest_d, latest_v = series[-1]
        return {
            "date": latest_d,
            "tonnes": latest_v,
            "change_5d": latest_v - series[-6][1],
        }
    except Exception as e_xlsx:  # noqa: BLE001
        try:
            data = json.loads(_get(GLD_JSON_URL).decode("utf-8", "replace"))
            node = data["data"]["total_tonnes"]
            return {
                "date": str(node.get("date", "?")),
                "tonnes": float(str(node["value"]).replace(",", "")),
                "change_5d": None,
            }
        except Exception as e:  # noqa: BLE001
            errors.append(
                f"GLD取得失敗: xlsx={type(e_xlsx).__name__}:{e_xlsx} / "
                f"json={type(e).__name__}:{e}"
            )
            return None


def fetch_calendar(errors: list[str]) -> list[dict] | None:
    """ForexFactory 今週の経済指標カレンダー（forecast/previousのみ、actualなし）。

    レート制限が厳しい（5分2回）ため、1日1回の朝実行でのみ叩くこと。
    """
    try:
        data = json.loads(_get(FF_CAL_URL).decode("utf-8", "replace"))
        events = []
        for ev in data:
            if ev.get("country") not in ("USD",):
                continue
            if ev.get("impact") not in ("High", "Medium"):
                continue
            try:
                dt = datetime.fromisoformat(ev["date"])
            except ValueError:
                continue
            events.append(
                {
                    "title": ev.get("title", "?"),
                    "impact": ev["impact"],
                    "utc": dt.astimezone(timezone.utc),
                    "forecast": ev.get("forecast", ""),
                    "previous": ev.get("previous", ""),
                }
            )
        return events
    except Exception as e:  # noqa: BLE001
        errors.append(f"カレンダー取得失敗: {type(e).__name__}: {e}")
        return None


def fetch_intraday(errors: list[str]) -> dict | None:
    """本日ここまでの日中足（5分足）。夜の「NY時間注意報」用。

    トレンド効率 = |正味の値動き| ÷ 総移動距離（1に近いほど一方向、
    0に近いほど往復）。H8検証前の参考値として扱うこと。
    """
    last_err: Exception | None = None
    for symbol in ("XAUUSD%3DX", "GC%3DF"):
        try:
            ch = _yahoo_chart(symbol, "1d", "5m")
            closes = ch["close"]
            if len(closes) < 12:
                raise ValueError(f"bars too few: {len(closes)}")
            o, c = ch["open"][0], closes[-1]
            path = sum(
                abs(closes[i] - closes[i - 1]) for i in range(1, len(closes))
            )
            return {
                "open": o,
                "close": c,
                "high": max(ch["high"]),
                "low": min(ch["low"]),
                "efficiency": (abs(c - o) / path) if path > 0 else None,
                "n_bars": len(closes),
                "symbol": "spot" if "XAUUSD" in symbol else "futures(GC=F)",
            }
        except Exception as e:  # noqa: BLE001
            last_err = e
    errors.append(f"日中足取得失敗: {type(last_err).__name__}: {last_err}")
    return None


def fetch_all(intraday: bool = False) -> dict:
    errors: list[str] = []
    out = {
        "gvz": fetch_gvz(errors),
        "price": fetch_price(errors),
        "cot": fetch_cot(errors),
        "gld": fetch_gld(errors),
        "calendar": fetch_calendar(errors),
        "errors": errors,
    }
    if intraday:
        out["intraday"] = fetch_intraday(errors)
    return out
